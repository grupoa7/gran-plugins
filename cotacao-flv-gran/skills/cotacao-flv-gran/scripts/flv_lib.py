"""
flv_lib.py — Biblioteca central da skill de Cotação Inteligente FLV (Gran).

Responsabilidades:
  - Carregar o dicionário de equivalência (ponte: descrição do fornecedor -> COD Gran)
  - Carregar a aba CONTAGEM FLV (demanda da rodada: COD, QTD, UND, fornecedor titular)
  - Normalizar preço para R$/kg ou R$/un
  - Banda de sanidade via último custo (pega erro de embalagem / preço absurdo)

Princípios (ver references/regras_negocio.md):
  - Dado real ou nada. Preço 0,00 / vazio = NÃO COTADO (nunca vence cherry-picking).
  - Cada ajuste toca UM arquivo. Conteúdo editável fica em templates/config.

Caminhos: a skill detecta a própria pasta via __file__, então funciona em host e sandbox.
"""
from __future__ import annotations
import os
import re
import unicodedata
from pathlib import Path

import openpyxl

SKILL_DIR = Path(__file__).resolve().parent.parent
# Dados mutáveis vivem no PROJETO (gravável), não no pacote da skill (read-only quando instalada).
# COTACAO_DADOS aponta para <projeto>/dados; fallback = dados/ ao lado do código (dev).
DADOS_DIR = Path(os.environ.get("COTACAO_DADOS") or (SKILL_DIR / "dados"))

# --------------------------------------------------------------------------- #
# QUADRO FIXO de fornecedores — fonte única em templates/fornecedores.json.
# DICIONARIO_FORNECEDORES (nome->coluna DESCRICAO no dicionário) e CANAL_BUSCA
# são DERIVADOS do registro, pra não ter que relembrar/ressincronizar a cada vez.
# Se o registro sumir/quebrar, cai no fallback hardcoded (sem regressão).
# --------------------------------------------------------------------------- #
import json as _json

FORNECEDORES_JSON = SKILL_DIR / "templates" / "fornecedores.json"


def carregar_fornecedores() -> list[dict]:
    """Lê o quadro fixo de fornecedores (roster + parâmetros). [] se ausente."""
    try:
        data = _json.loads(FORNECEDORES_JSON.read_text(encoding="utf-8"))
        return data.get("fornecedores", [])
    except Exception:
        return []


_REGISTRO = carregar_fornecedores()
if _REGISTRO:
    DICIONARIO_FORNECEDORES = {f["nome"]: f["col_dict"] for f in _REGISTRO}
    CANAL_BUSCA = {f["nome"] for f in _REGISTRO if f.get("canal") == "busca"}
    # Micael/Rota são busca histórica sem tabela cadastrada — mantém p/ titular.
    CANAL_BUSCA |= {"MICAEL", "ROTA (MICAEL)"}
else:  # fallback se o registro sumir (mantém o comportamento antigo)
    DICIONARIO_FORNECEDORES = {
        "SHIMIZU": 5, "RML": 9, "DOCE MEL": 13, "DONOFRIO": 17,
        "HORTIMIX": 21, "BOA CITRUS": 25, "IGARASHI": 29,
    }
    CANAL_BUSCA = {"RML", "MICAEL", "ROTA (MICAEL)", "HORTIMIX", "BOA CITRUS"}


# --------------------------------------------------------------------------- #
# Normalização de texto para casamento de descrições
# --------------------------------------------------------------------------- #
def norm_texto(s: str) -> str:
    """Minúsculas, sem acento, sem pontuação, espaços colapsados. Base do match."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # unifica espaçamento de unidade pra casar "20kg"=="20 kg", "10x800g"=="10 x 800 g"
    s = re.sub(r"(\d)\s+(kg|g|x|un|und|uni)\b", r"\1\2", s)
    s = re.sub(r"\b(cx|sc|saco)\s+(\d)", r"\1\2", s)
    return s


# --------------------------------------------------------------------------- #
# Preço: parsing robusto de número em pt-BR
# --------------------------------------------------------------------------- #
def parse_preco(valor) -> float | None:
    """'R$ 1.234,56' / '12,00' / 12.0 -> float. Retorna None se vazio/0/zerado."""
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        v = float(valor)
        return v if v > 0 else None
    txt = str(valor)
    txt = re.sub(r"(?i)r\$|\s", "", txt)
    if not txt:
        return None
    # remove milhar (.) e troca decimal (,) por ponto
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    try:
        v = float(txt)
    except ValueError:
        return None
    return v if v > 0 else None


# --------------------------------------------------------------------------- #
# Unidade -> peso em kg (para normalizar caixa/saca em R$/kg)
# --------------------------------------------------------------------------- #
_RE_KG = re.compile(r"(\d+[.,]?\d*)\s*kg", re.I)
_RE_NXG = re.compile(r"(\d+)\s*x\s*(\d+[.,]?\d*)\s*g", re.I)   # 10x800g
_RE_G = re.compile(r"(\d+[.,]?\d*)\s*g\b", re.I)


def peso_kg_da_unidade(unidade: str) -> float | None:
    """Extrai o peso em kg de uma string de embalagem. None se não der pra inferir.
    Ex: 'CX 20KG'->20 ; 'SC 50KG'->50 ; '10X800G'->8.0 ; '250G'->0.25 ; 'KG'->1 ; 'UND'->None
    """
    if not unidade:
        return None
    u = str(unidade).strip()
    un = norm_texto(u)
    if un in ("kg", "quilo", "kilo"):
        return 1.0
    m = _RE_NXG.search(u)
    if m:
        return int(m.group(1)) * float(m.group(2).replace(",", ".")) / 1000.0
    m = _RE_KG.search(u)
    if m:
        return float(m.group(1).replace(",", "."))
    m = _RE_G.search(u)
    if m:
        return float(m.group(1).replace(",", ".")) / 1000.0
    return None  # unidade por peça (UND/CENTO/MOL/etc) -> não converte pra kg


# --------------------------------------------------------------------------- #
# Base de embalagens (caixa padrão CEASA-BA) — fallback quando a string da
# unidade do fornecedor NÃO traz o peso. Ex.: HR/Igarashi/Boa Citrus/Qualisuper/
# Potência mandam "tomate caixa R$ 120" sem declarar kg. A função abaixo
# consulta templates/embalagens.json (gerado da Tabela_Conversao_Caixas_FLV_Gran)
# e devolve peso default + flag nao_padrao (true bloqueia conversão automática).
# --------------------------------------------------------------------------- #
EMBALAGENS_JSON = SKILL_DIR / "templates" / "embalagens.json"


def carregar_embalagens() -> dict:
    """Carrega a base de embalagens (peso por commodity + mapping SKU->commodity).
    Estrutura ver templates/embalagens.json. Cache simples evita reler a cada call.
    """
    if not hasattr(carregar_embalagens, "_cache"):
        if EMBALAGENS_JSON.exists():
            with open(EMBALAGENS_JSON, encoding="utf-8") as fp:
                carregar_embalagens._cache = _json.load(fp)
        else:
            carregar_embalagens._cache = {"produtos": {}, "skus": {}}
    return carregar_embalagens._cache


def peso_caixa_por_commodity(commodity_slug: str) -> tuple[float | None, bool, list[str]]:
    """Consulta peso padrão CEASA-BA por commodity. Retorna (peso_kg, nao_padrao, motivo).
    - peso_kg=None: produto sem peso convertível (cento/unidade/maço/kg)
    - nao_padrao=True: NÃO converter automaticamente; exigir peso do fornecedor.
    Ex: 'tomate_extra_primeira_salada_italiano' -> (21.0, False, [])
    Ex: 'banana_prata' -> (45.0, True, ['peso_fora_padrao_universal_45kg', ...])
    """
    if not commodity_slug:
        return None, True, ["commodity_nao_mapeada"]
    base = carregar_embalagens()
    p = base.get("produtos", {}).get(commodity_slug)
    if not p:
        return None, True, ["commodity_nao_cadastrada"]
    kg = p.get("kg_por_caixa")
    return (float(kg) if kg is not None else None), bool(p.get("nao_padrao")), list(p.get("motivo_nao_padrao") or [])


def peso_caixa_por_sku(cod_gran: int | str) -> tuple[float | None, bool, list[str], str | None]:
    """Consulta peso padrão por COD Gran (atalho que devolve também a commodity).
    Retorna (peso_kg, nao_padrao, motivo, commodity_slug).
    """
    base = carregar_embalagens()
    sku = base.get("skus", {}).get(str(cod_gran))
    if not sku:
        return None, True, ["sku_nao_cadastrado"], None
    slug = sku.get("commodity_slug")
    if sku.get("pendente_confirmacao"):
        kg = sku.get("peso_kg_default")
        return (float(kg) if kg is not None else None), True, ["pendente_confirmacao"], slug
    return (*peso_caixa_por_commodity(slug), slug)


# Strings de embalagem que JÁ trazem o peso explícito (ou são por peça/contagem).
# Se a unidade casa com algum desses padrões, NÃO consultamos embalagens.json — o
# fornecedor declarou. A consulta à base só rola pra caixas/sacas sem peso embutido.
_RE_PESO_EXPLICITO = re.compile(
    r"\b(\d+[.,]?\d*\s*(?:kg|g)|\d+\s*x\s*\d+\s*g)\b", re.I)
_RE_UNID_PECA = re.compile(
    r"\b(kg|quilo|kilo|und?|uni|unid|unidade|pc|pct|pacote|mol|maco|maço|cento|"
    r"duzia|duzias|dz|bdj|bandeja)\b", re.I)
_RE_CAIXA_SACA = re.compile(r"\b(cx|caixa|sc|saco|saca|fardo|frd|fd)\b", re.I)


def resolver_peso_kg(unidade: str | None, cod_gran: int | str | None = None) -> dict:
    """Resolve o peso em kg que normaliza preço/caixa pra R$/kg.

    REGRA DE OURO (validada com Hugo 2026-05-28):
      O fornecedor é a fonte primária. Só caímos pra embalagens.json (CEASA-BA)
      quando a string da unidade NÃO traz peso E NÃO é por peça/kg/unidade.
      Nunca sobrescreve declaração explícita do fornecedor.

    Devolve dict com:
      - peso_kg:      float | None   (None = não dá pra normalizar pra kg)
      - origem:       'explicito_fornecedor' | 'ceasa_base' | 'por_peca' | 'nao_resolvido'
      - nao_padrao:   bool           (True quando origem=ceasa_base mas commodity é exceção
                                      — sistema deve BLOQUEAR conversão e exigir kg do fornec.)
      - motivo:       list[str]      (rastreio: por que esse peso e essa flag)
      - commodity:    str | None     (slug consultado, se houve fallback)
    """
    u = (unidade or "").strip()

    # 1) Peso EXPLÍCITO na string ("CX 20KG", "10x800G", "250G") -> autoridade do fornecedor
    peso = peso_kg_da_unidade(u)
    if peso is not None and _RE_PESO_EXPLICITO.search(u):
        return {"peso_kg": peso, "origem": "explicito_fornecedor",
                "nao_padrao": False, "motivo": ["peso_na_string_unidade"], "commodity": None}

    # 2) Unidade é KG puro (preço já por kg) -> peso=1
    if u and re.fullmatch(r"\s*(kg|quilo|kilo)\s*", u, re.I):
        return {"peso_kg": 1.0, "origem": "explicito_fornecedor",
                "nao_padrao": False, "motivo": ["unidade_kg_direta"], "commodity": None}

    # 3) Unidade por peça (UND/MOL/CENTO/BDJ/etc) -> peso None, não converte pra kg
    if u and _RE_UNID_PECA.search(u) and not _RE_CAIXA_SACA.search(u):
        return {"peso_kg": None, "origem": "por_peca",
                "nao_padrao": False, "motivo": ["unidade_por_peca_ou_contagem"],
                "commodity": None}

    # 4) Sem peso explícito E parece caixa/saca (ou vazio) -> fallback embalagens.json
    if cod_gran is not None:
        peso_b, np_, motivo_b, slug = peso_caixa_por_sku(cod_gran)
        if peso_b is not None:
            return {"peso_kg": peso_b, "origem": "ceasa_base",
                    "nao_padrao": np_,
                    "motivo": ["fallback_ceasa_base", *motivo_b],
                    "commodity": slug}
        # SKU sem peso na base (commodity vendida por unidade/cento)
        return {"peso_kg": None, "origem": "nao_resolvido",
                "nao_padrao": True,
                "motivo": ["sku_sem_peso_na_base", *motivo_b],
                "commodity": slug}

    # 5) Sem cod e sem peso -> não dá pra resolver
    return {"peso_kg": None, "origem": "nao_resolvido",
            "nao_padrao": True, "motivo": ["sem_cod_e_sem_peso"], "commodity": None}


def registrar_pesagem(commodity_slug: str, peso_kg: float, sku_cod: int | str | None = None,
                       fornecedor: str | None = None, data: str | None = None) -> dict:
    """Anexa uma pesagem REAL (recebimento do Gran) em produtos[slug].pesagens.
    Aplica regra de calibração: se ≥2 pesagens divergem >threshold% do default,
    devolve {'sugerir_troca': True, 'novo_default': <mediana>} pra Hugo aprovar.
    Persiste no embalagens.json.
    """
    import statistics, datetime
    base = carregar_embalagens()
    prod = base.get("produtos", {}).get(commodity_slug)
    if not prod:
        return {"ok": False, "erro": f"commodity '{commodity_slug}' não encontrada"}
    default = prod.get("kg_por_caixa")
    pesagem = {
        "data": data or datetime.date.today().isoformat(),
        "peso_kg": round(float(peso_kg), 2),
        "sku_cod": str(sku_cod) if sku_cod is not None else None,
        "fornecedor": fornecedor,
    }
    prod.setdefault("pesagens", []).append(pesagem)
    # Calibração
    regras = base.get("regras_calibracao", {})
    threshold_pct = regras.get("threshold_divergencia_pct", 10)
    minimo = regras.get("pesagens_minimas_para_troca", 2)
    sugerir = None
    if default and len(prod["pesagens"]) >= minimo:
        divergentes = [p["peso_kg"] for p in prod["pesagens"]
                       if abs(p["peso_kg"] - default) / default * 100 > threshold_pct]
        if len(divergentes) >= minimo:
            sugerir = round(statistics.median(divergentes), 2)
    # Persiste
    with open(EMBALAGENS_JSON, "w", encoding="utf-8") as fp:
        _json.dump(base, fp, ensure_ascii=False, indent=2)
    # Invalida cache
    if hasattr(carregar_embalagens, "_cache"):
        del carregar_embalagens._cache
    return {
        "ok": True,
        "default_atual": default,
        "pesagem_registrada": pesagem,
        "total_pesagens": len(prod["pesagens"]),
        "sugerir_troca": sugerir is not None,
        "novo_default_sugerido": sugerir,
    }


# --------------------------------------------------------------------------- #
# Carregar dicionário de equivalência
# --------------------------------------------------------------------------- #
def carregar_dicionario(path: Path | None = None) -> dict:
    """Retorna estrutura:
      {
        'por_cod':   {cod: {'desc','grupo','subgrupo','unidade'}},
        'match':     {FORNECEDOR: {norm_desc: cod}},
        'desc_forn': {FORNECEDOR: {cod: (descricao_original, unidade_forn)}},
      }
    """
    # Versão oficial revisada pela compradora (substitui a anterior, ignorada).
    path = path or (DADOS_DIR / "dicionario_equivalencia_oficial.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    header = rows[0]
    por_cod, match, desc_forn = {}, {}, {}
    for f in DICIONARIO_FORNECEDORES:
        match[f] = {}
        desc_forn[f] = {}
    for r in rows[1:]:
        cod = r[0]
        if cod is None:
            continue
        por_cod[cod] = {
            "desc": r[3], "grupo": r[1], "subgrupo": r[2], "unidade": r[4],
        }
        for f, di in DICIONARIO_FORNECEDORES.items():
            # Guard defensivo: fornecedor cadastrado no roster mas sem colunas no XLSX
            # ainda. Pula sem quebrar a rodada (caso típico: fornecedor novo só com peso/
            # entrada parcial). Vira nao_casados; o roster avisa no stderr.
            if di + 1 >= len(r):
                continue
            d = r[di]
            if d:
                # uma célula pode listar vários descritores p/ o mesmo COD separados por '|'
                # (ex.: Igarashi plano A | plano B). Cada um vira chave de match -> mesmo COD.
                for part in str(d).split("|"):
                    nd = norm_texto(part)
                    if nd:
                        match[f][nd] = cod
                desc_forn[f][cod] = (d, r[di + 1])  # descricao, unidade
    return {"por_cod": por_cod, "match": match, "desc_forn": desc_forn}


# --------------------------------------------------------------------------- #
# Carregar aba CONTAGEM FLV (demanda da rodada)
# --------------------------------------------------------------------------- #
# Índices (cabeçalho real na linha 6, idx 5): COD[1] DESC[4] UND[5] QTD[18] FORN[17] CUSTO[19]
CONT_COL = {"cod": 1, "desc": 4, "und": 5, "qtd": 18, "forn": 17, "custo": 19, "curva": 14}


def carregar_contagem(path: Path, somente_a_pedir: bool = True) -> list[dict]:
    """Lê a aba CONTAGEM FLV. Retorna itens da rodada (QTD>0 por padrão)."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["CONTAGEM FLV"] if "CONTAGEM FLV" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # localizar a linha de cabeçalho (onde col1 == 'COD')
    hdr_idx = next((i for i, r in enumerate(rows[:15]) if str(r[1]).strip().upper() == "COD"), 5)
    itens = []
    for r in rows[hdr_idx + 1:]:
        cod = r[CONT_COL["cod"]]
        if cod is None or not str(cod).strip().isdigit():
            continue
        qtd = r[CONT_COL["qtd"]]
        try:
            qtd = float(qtd) if qtd is not None else 0.0
        except (ValueError, TypeError):
            qtd = 0.0
        if somente_a_pedir and qtd <= 0:
            continue
        custo = r[CONT_COL["custo"]]
        try:
            custo = float(custo) if custo not in (None, "") else None
        except (ValueError, TypeError):
            custo = None
        itens.append({
            "cod": int(cod),
            "desc": r[CONT_COL["desc"]],
            "und": r[CONT_COL["und"]],
            "qtd": qtd,
            "fornecedor_titular": (r[CONT_COL["forn"]] or "").strip(),
            "custo_titular_total": custo,
            "curva": (r[CONT_COL["curva"]] or "").strip(),
        })
    return itens


# --------------------------------------------------------------------------- #
# Banda de sanidade
# --------------------------------------------------------------------------- #
def fora_da_banda(preco_kg: float, ultimo_custo_kg: float | None, fator: float = 2.0) -> bool:
    """True se o preço cotado destoa absurdamente do último custo (> fator x ou < 1/fator)."""
    if not preco_kg or not ultimo_custo_kg or ultimo_custo_kg <= 0:
        return False
    razao = preco_kg / ultimo_custo_kg
    return razao > fator or razao < (1.0 / fator)


def carregar_vendas(path: Path | None = None) -> dict:
    """Lê vendas_60d (aba PRECIFICAÇÃO) p/ contexto de giro nos gráficos.
    Retorna {cod: {desc, grupo, qtd, fat, curva}}."""
    path = path or (DADOS_DIR / "vendas_60d.xlsx")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    out = {}
    for r in rows[2:]:
        cod = r[0]
        if cod is None:
            continue
        def n(i):
            try:
                return float(r[i]) if r[i] not in (None, "") else 0.0
            except (ValueError, TypeError):
                return 0.0
        out[cod] = {"desc": r[4], "grupo": (r[3] or "").strip(),
                    "qtd": n(10), "fat": n(11), "curva": (r[16] or "").strip(),
                    "perda_qtd": n(7), "perda_valor": n(8), "perda_pct": n(9),
                    "preco_venda": n(55), "custo_unit": n(50), "margem_pct": n(54)}
    return out


# --------------------------------------------------------------------------- #
# BI do Gran (aba APOIO PEDIDO) — base oficial da DEMANDA da cotação
# --------------------------------------------------------------------------- #
# Decisão 25/05: a base de venda de cada produto é a MÉDIA FINAL (col BA), que é
# a média DIÁRIA ponderada (exclui outliers, considera novos produtos). O pedido
# semanal NÃO é mais a base de demanda. Giro semanal = MÉDIA FINAL × dias/semana.
# Cabeçalho na linha 2; dados a partir da linha 3. Índices 0-based:
BI_COL = {"cod": 0, "grupo": 1, "subgrupo": 2, "desc": 3, "unid": 4, "curva": 17,
          "pcusto": 39, "patual": 41, "forn": 47, "media_final": 52}
BI_DIAS_SEMANA = 7  # Gran abre todo dia


def carregar_bi(path, dias_semana: int = BI_DIAS_SEMANA) -> dict:
    """Lê a aba 'APOIO PEDIDO' do BI. Retorna {cod: {desc, unid, curva,
    custo_unit (P.CUSTO), preco_venda (P.ATUAL), forn (FORNECEDOR atual),
    media_final (BA, diária), giro_sem (= media_final × dias_semana)}}."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["APOIO PEDIDO"] if "APOIO PEDIDO" in wb.sheetnames else wb.active

    def num(v):
        try:
            return float(v) if v not in (None, "") else 0.0
        except (ValueError, TypeError):
            return 0.0

    out = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        if len(r) <= BI_COL["media_final"]:
            continue
        cod = r[BI_COL["cod"]]
        if cod is None:
            continue
        try:
            cod = int(cod)
        except (ValueError, TypeError):
            continue
        media = num(r[BI_COL["media_final"]])
        out[cod] = {
            "desc": r[BI_COL["desc"]],
            "unid": r[BI_COL["unid"]] or "",
            "curva": str(r[BI_COL["curva"]] or "").strip(),
            "custo_unit": num(r[BI_COL["pcusto"]]),
            "preco_venda": num(r[BI_COL["patual"]]),
            "forn": str(r[BI_COL["forn"]] or "").strip(),
            "media_final": media,
            "giro_sem": round(media * dias_semana, 3),
        }
    wb.close()
    return out


def demanda_do_bi(bi: dict, dic: dict, titular_pedido: dict | None = None,
                  somente_com_giro: bool = True) -> tuple[list, dict]:
    """Monta (contagem, vendas) a partir do BI. Universo = SKUs do cadastro
    (dicionário) presentes no BI COM demanda (MÉDIA FINAL > 0). Itens de giro zero
    são excluídos por padrão (poluem o Mapa à toa — decisão Hugo 25/05); passe
    somente_com_giro=False para incluir todos. `titular_pedido` {cod: nome_forn}
    (Pedido FLV) sobrepõe o fornecedor do BI quando disponível (titular mais atual)."""
    titular_pedido = titular_pedido or {}
    contagem, vendas = [], {}
    for cod, info in dic["por_cod"].items():
        b = bi.get(cod)
        if not b:
            continue
        giro = b["giro_sem"]
        if somente_com_giro and giro <= 0:
            continue            # sem demanda no BI -> fora da cotação
        titular = (titular_pedido.get(cod) or b["forn"] or "").strip()
        contagem.append({
            "cod": cod,
            "desc": b["desc"] or info.get("desc"),
            "und": info.get("unidade") or b["unid"],
            "qtd": giro,                       # giro semanal (base = BI)
            "fornecedor_titular": titular,
            "custo_titular_total": (b["custo_unit"] * giro) if giro else None,
            "curva": b["curva"],
        })
        vendas[cod] = {"desc": b["desc"], "preco_venda": b["preco_venda"],
                       "custo_unit": b["custo_unit"], "curva": b["curva"],
                       "giro_sem": giro, "qtd": 0.0}
    return contagem, vendas


if __name__ == "__main__":
    d = carregar_dicionario()
    print("Dicionário:", len(d["por_cod"]), "SKUs")
    for f in DICIONARIO_FORNECEDORES:
        print(f"  {f}: {len(d['match'][f])} descrições mapeadas")
    print("\nTeste peso_kg_da_unidade:")
    for u in ["CX 20KG", "SC 50KG", "10X800G", "250G", "KG", "UND", "CENTO", "CX 9 KG"]:
        print(f"  {u!r} -> {peso_kg_da_unidade(u)}")
    print("\nTeste parse_preco:")
    for p in ["R$ 1.234,56", "12,00", "0,00", 0, 7.0, "", None]:
        print(f"  {p!r} -> {parse_preco(p)}")
