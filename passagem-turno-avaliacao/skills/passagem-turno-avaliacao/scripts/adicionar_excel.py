#!/usr/bin/env python3
"""
adicionar_excel.py — Adiciona linha de avaliação no XLSX acumulativo.

ESTRUTURA DA PLANILHA (a partir de 18/05/2026):
- Aba "Avaliações": fonte única da verdade. Colunas A-I = dados manuais; J-N = fórmulas.
- Aba "Notas (long)": unpivot automático via fórmula (NÃO MEXER).
- Aba "Médias": fórmulas que calculam média móvel últimas 7 por encarregado.
- Aba "Cenário": thresholds editáveis (±5pp, ±15pp).
- Aba "Resumo Semanal": agregação por janela qua→ter.
- Aba "Leia-me": instruções.

Este script SÓ adiciona linha em "Avaliações" (cols A-I + obs em N). Fórmulas das outras abas se atualizam sozinhas ao abrir no Excel.

Uso:
    # Adicionar nova linha
    python adicionar_excel.py --card-id 1351778171 \
        --data-card "17/05/2026" \
        --abertura-nome "Sílvio Rouzan" --abertura-pct 100 \
        --tarde-nome "Sílvio Rouzan" --tarde-pct 98 \
        --loja-pct 99 \
        --observacao "Turno único dom; Alyne pulada (folga padrão)"

    # Ler histórico
    python adicionar_excel.py --ler-historico --periodo 30
"""

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path


def detectar_workspace():
    """Detecta o path do workspace do projeto.

    1. Tentar o path absoluto do mac do Hugo.
    2. Fallback: Path(__file__).parent.parent.parent (skill mora dentro do workspace).
    """
    mac_path = Path("/Users/hugogusmao/Documents/Claude/Projects/[GRAN] Passagem de Turno")
    try:
        if mac_path.exists():
            return mac_path
    except (PermissionError, OSError):
        pass
    return Path(__file__).resolve().parent.parent.parent


XLSX_PATH = detectar_workspace() / "passagem-turno-superGran.xlsx"

# Cabeçalho da aba "Avaliações" — fonte única da verdade.
# Colunas J-N são preenchidas por fórmula automaticamente.
HEADER_AVALIACOES = [
    "Data Card", "Card ID", "Encarregado Abertura", "Nota Abertura (%)",
    "Encarregado Tarde/Noite", "Nota Tarde/Noite (%)", "Nota Loja (%)",
    "Avaliado em", "Link do Card",
    "Cenário Abertura", "Δ Abertura (pp)",
    "Cenário Tarde/Noite", "Δ Tarde/Noite (pp)",
    "Observação",
]

SHEET_AVALIACOES = "Avaliações"


def get_workbook():
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print("ERRO: openpyxl não instalado. Rode: pip install openpyxl --break-system-packages", file=sys.stderr)
        sys.exit(1)

    if XLSX_PATH.exists():
        wb = load_workbook(XLSX_PATH)
        if SHEET_AVALIACOES not in wb.sheetnames:
            print(f"ERRO: aba '{SHEET_AVALIACOES}' não existe. Estrutura quebrada.", file=sys.stderr)
            sys.exit(2)
        return wb, wb[SHEET_AVALIACOES]

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_AVALIACOES
    ws.append(HEADER_AVALIACOES)
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        cell.fill = PatternFill(start_color="37474F", end_color="37474F", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for i, w in enumerate([12,12,22,14,22,16,12,18,36,14,14,16,14,36], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return wb, ws


def encontrar_proxima_linha(ws):
    """Primeira linha vazia (procurando coluna B = Card ID)."""
    for r in range(2, 502):
        if ws.cell(row=r, column=2).value in (None, ""):
            return r
    print("ERRO: planilha cheia (500 linhas).", file=sys.stderr)
    sys.exit(3)


def adicionar_linha(args):
    wb, ws = get_workbook()
    link = f"https://app.pipefy.com/open-cards/{args.card_id}"
    avaliado_em = args.avaliado_em or datetime.now().strftime("%d/%m/%Y %H:%M")
    try:
        data_dt = datetime.strptime(args.data_card, "%d/%m/%Y")
    except ValueError:
        print(f"ERRO: data-card '{args.data_card}' não está em DD/MM/AAAA", file=sys.stderr)
        sys.exit(4)

    row_idx = encontrar_proxima_linha(ws)
    ws.cell(row=row_idx, column=1, value=data_dt).number_format = "DD/MM/YYYY"
    ws.cell(row=row_idx, column=2, value=str(args.card_id))
    ws.cell(row=row_idx, column=3, value=args.abertura_nome)
    ws.cell(row=row_idx, column=4, value=float(args.abertura_pct))
    ws.cell(row=row_idx, column=5, value=args.tarde_nome or args.abertura_nome)
    if args.tarde_pct is not None:
        ws.cell(row=row_idx, column=6, value=float(args.tarde_pct))
    ws.cell(row=row_idx, column=7, value=float(args.loja_pct))
    ws.cell(row=row_idx, column=8, value=avaliado_em)
    ws.cell(row=row_idx, column=9, value=link)
    if args.observacao:
        ws.cell(row=row_idx, column=14, value=args.observacao)

    wb.save(XLSX_PATH)
    print(json.dumps({
        "status": "ok", "arquivo": str(XLSX_PATH), "aba": SHEET_AVALIACOES,
        "linha_adicionada": row_idx, "card_id": args.card_id, "data_card": args.data_card,
        "nota_loja": args.loja_pct, "nota_abertura": args.abertura_pct, "nota_tarde": args.tarde_pct,
        "observacao": args.observacao,
        "nota": "Cols J-M (Cenário/Δ) são fórmulas — recalcular ao abrir no Excel. Aba 'Médias' atualiza sozinha."
    }, ensure_ascii=False, indent=2))


def ler_historico(periodo_dias):
    if not XLSX_PATH.exists():
        print(json.dumps({"status": "vazio", "msg": "XLSX não existe."}, ensure_ascii=False))
        return
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_PATH, data_only=True)
    if SHEET_AVALIACOES not in wb.sheetnames:
        print(json.dumps({"status": "erro", "msg": "Aba 'Avaliações' ausente."}, ensure_ascii=False))
        return
    ws = wb[SHEET_AVALIACOES]
    rows = []
    for r in ws.iter_rows(min_row=2, max_row=200, values_only=True):
        if r[1]:
            rows.append(r)
    if not rows:
        print(json.dumps({"status": "vazio", "msg": "Aba Avaliações sem registros."}, ensure_ascii=False))
        return
    corte = datetime.now() - timedelta(days=periodo_dias)
    filtradas = []
    for r in rows:
        d = r[0]
        if isinstance(d, datetime) and d >= corte:
            filtradas.append(r)
        elif isinstance(d, str):
            try:
                if datetime.strptime(d, "%d/%m/%Y") >= corte:
                    filtradas.append(r)
            except ValueError:
                pass
    if not filtradas:
        print(json.dumps({"status": "vazio", "msg": f"Sem registros nos últimos {periodo_dias} dias."}, ensure_ascii=False))
        return
    notas_loja = [float(r[6]) for r in filtradas if r[6] is not None]
    media_loja = round(sum(notas_loja) / len(notas_loja), 1) if notas_loja else 0
    encarregados = {}
    for r in filtradas:
        for nome_idx, pct_idx in [(2, 3), (4, 5)]:
            nome, pct = r[nome_idx], r[pct_idx]
            if nome and pct is not None:
                encarregados.setdefault(nome, []).append(float(pct))
    enc_stats = {n: {"cards": len(p), "media_pct": round(sum(p)/len(p),1), "min": min(p), "max": max(p)} for n,p in encarregados.items()}
    tendencia = "indefinida"
    if len(notas_loja) >= 4:
        meio = len(notas_loja) // 2
        delta = sum(notas_loja[meio:])/(len(notas_loja)-meio) - sum(notas_loja[:meio])/meio
        tendencia = "subindo" if delta > 5 else "caindo" if delta < -5 else "estavel"
    fmt = lambda d: d.strftime("%d/%m/%Y") if isinstance(d, datetime) else str(d)
    print(json.dumps({
        "status": "ok", "periodo_dias": periodo_dias, "total_cards": len(filtradas),
        "media_loja_pct": media_loja, "tendencia": tendencia, "encarregados": enc_stats,
        "ultimos_5_cards": [{
            "data": fmt(r[0]), "card_id": r[1],
            "abertura": r[2], "abertura_pct": r[3],
            "tarde": r[4], "tarde_pct": r[5],
            "loja_pct": r[6], "cenario_ab": r[9], "delta_ab": r[10],
            "cenario_tn": r[11], "delta_tn": r[12], "observacao": r[13], "link": r[8]
        } for r in filtradas[-5:]]
    }, ensure_ascii=False, indent=2))


def main():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--card-id")
    p.add_argument("--data-card")
    p.add_argument("--abertura-nome")
    p.add_argument("--abertura-pct", type=float)
    p.add_argument("--tarde-nome")
    p.add_argument("--tarde-pct", type=float)
    p.add_argument("--loja-pct", type=float)
    p.add_argument("--avaliado-em")
    p.add_argument("--observacao")
    p.add_argument("--ler-historico", action="store_true")
    p.add_argument("--periodo", type=int, default=30)
    args = p.parse_args()

    if args.ler_historico:
        ler_historico(args.periodo)
    else:
        required = ["card_id", "data_card", "abertura_nome", "abertura_pct", "loja_pct"]
        missing = [r for r in required if getattr(args, r) is None]
        if missing:
            print(f"ERRO: faltando argumentos: {missing}", file=sys.stderr)
            sys.exit(1)
        adicionar_linha(args)


if __name__ == "__main__":
    main()
